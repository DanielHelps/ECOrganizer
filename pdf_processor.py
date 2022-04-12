from pdfminer.layout import LAParams, LTTextBox
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
import openCV_circles
from openpyxl import load_workbook
import re
import os
import math
import win32com.client as win32
import tkinter as tk

# Create a page object to scan that contains the parameters and pages for pdf scanning
def get_pages_obj(path):
    fp = open(path, 'rb')
    rsrcmgr = PDFResourceManager()
    # laparams = LAParams(line_overlap=0.1, char_margin=0.1, line_margin=0.01)
    laparams = LAParams(line_overlap=0.1, char_margin=0.0000000001, word_margin=0.00000000001, line_margin=0.01)

    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    # Getting pages objects (1 object per page)
    pages = PDFPage.get_pages(fp)

    # Creating a layout object that contains all curves, text boxes and shapes for each page
    layout = []
    for page in pages:
        interpreter.process_page(page)
        layout.append(device.get_result())
    return layout


# Function to search within a text box for a P/N in the form of XX-YYYY-XXXX where X
# is a number and Y is an uppercase letter
def get_pn(text, current_pn):
    check_for_pn = re.findall('\\d{2}-[A-Z]{4}-\\d{4}', text)
    # If there are multiple kornit P/N's, return only the last
    # one in the text box (as this one is the correct one according to tests)
    if len(check_for_pn) != 0:
        return check_for_pn[len(check_for_pn) - 1]
    # If only one P/N
    else:
        return current_pn


# Function to search within a text box for a drawing number in the form of XXX-XX-XX-XXX where X is a number
def get_drawing_number(text, current_drawing_number):
    # Find all occurrences of a drawing number in the text box
    check_for_drawing_number = re.findall('\\d{3}-\\d{2}-\\d{2}-\\d{3}', text)
    # If there are multiple drawing numbers, return only the last
    # one in the text box (as this one is the correct one according to tests)
    if len(check_for_drawing_number) != 0:
        return check_for_drawing_number[len(check_for_drawing_number) - 1]
    else:
        return current_drawing_number


# Function to search for the revision of the drawing within the PDF
def get_rev(text, y_try, current_y, current_rev):
    # Search for a string with the revision AX or BX where X is a number from 0 to 9
    rev_try = re.findall('[A-B]\\d', text)
    # If exists a AX or BX within the text box and the y coordinate is larger than
    # the current candidate for the revision then choose the higher y coordinate revision
    # This occurs because sheet size also has the same form, but it has lower y coordinate
    if rev_try and y_try > current_y:
        return rev_try, y_try
    # If multiple strings of AX or BX exist in the text box, take the first one
    elif rev_try and len(rev_try) > 1:
        return rev_try[0], y_try
    else:
        return current_rev, current_y


# Function to search for signatures in PDF
def check_signatures(text, x, y, drawn_x, drawn_y, number_of_signatures):
    # check that the current tested text is just to the right (and maybe a little below) of
    # "Drawn by" text and has letters in it (not '-')
    if drawn_x + 0.01 < x < drawn_x + 0.05 and y < drawn_y + 0.01 and re.search('[a-zA-Z]', text) != None:
        processed_signatures = re.findall('(.*)\n', text)
        # Text box can have multiple signatures
        for signature in processed_signatures:
            # Check that signature has more than 2 characters (1 or 2 characters is
            # considered to not be a signature)
            if len(signature.strip()) > 2:
                number_of_signatures += 1
        # print(number_of_signatures)
        # print("YES")
        return number_of_signatures
    else:
        # print("n")
        return number_of_signatures


# Function to check that date exists next to signature (at least 1 signature required)
def check_date(text, x, y):
    # location of date will be between 70% and 80% of sheet to the right
    # and less than 10% of sheet upwards, also it can't have the character '-' in it (. or / are allowed)
    # if 0.64 < x < 0.8 and y < 0.1 and re.search('\d*\d*\d\d', text.strip()) != None and re.search('-', text.strip()) == None and re.search('[a-zA-Z]', text.strip()) == None:
    if 0.64 < x < 0.8 and y < 0.1 and re.search('\d*\d*\d\d', text.strip()) != None and re.search('-', text.strip()) == None:
        # Return True if a signature exists in this text box, otherwise False
        return True
    else:
        return False


# A function to count the amount of balloons required in an assembly (for xls document)
def xls_max_balloons(path, digits):
    # save as xlsx document
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(path)
    wb.SaveAs(path + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension

    # load workbook
    wrkbk = load_workbook(f"{path[:-digits]}.xlsx")
    sh = wrkbk.active
    # amount of balloons = amount of rows in BOM table
    max_balloons = sh.max_row - 1
    excel.Application.Quit()
    # Delete xlsx file (keep the xls)
    os.remove(path + "x")
    return max_balloons


# A function to count the amount of balloons required in an assembly (for any Excel document)
def number_of_balloons(path):
    os.getcwd()
    # if path like 964-19-00-000_A0.xlsx exists
    if os.path.exists(f"{path[:-7]}.xlsx"):
        wrkbk = load_workbook(f"{path[:-7]}.xlsx")
        sh = wrkbk.active
        max_balloons = sh.max_row - 1
    # if path like 964-19-00-000.xlsx exists
    elif os.path.exists(f"{path[:-4]}.xlsx"):
        wrkbk = load_workbook(f"{path[:-4]}.xlsx")
        sh = wrkbk.active
        max_balloons = sh.max_row - 1
    # if path like 964-19-00-000_A0.xls exists
    elif os.path.exists(f"{path[:-7]}.xls"):
        max_balloons = xls_max_balloons(f"{os.getcwd()}\\{path[:-7]}.xls", 7)
    # if path like 964-19-00-000.xls exists
    elif os.path.exists(f"{path[:-4]}.xls"):
        max_balloons = xls_max_balloons(f"{os.getcwd()}\\{path[:-4]}.xls", 4)
    # If no excel exists - not an assembly
    else:
        max_balloons = 0
    return max_balloons

    #


# A function that counts the missing balloons in an assembly
def check_missing_circles(x, y, centers_x, centers_y, text):
    # Centers of the circles detected by openCV.HoughCircles
    centers = list(zip(centers_x, centers_y))
    for a in centers:
        # Check distance of location of text scanned by pdf
        # from the center of each openCV.HoughCircles detected circles
        distance = math.dist([x, y], a)
        # Center of text and detected circle needs to be less than 1.5% of sheet size apart,
        # and text must satisfy that it is a number only (one or double digit)
        if distance < 0.015 and (re.fullmatch('\\d\\d\n', text) is not None or re.fullmatch('\\d\n', text) is not None):
            return int(text)
    # If text is not a circle
    return None


def check_again_signatures(layout, page_height, page_width, drawn_x, drawn_y):
    numbers_of_signatures = 0
    for obj in layout:
        if isinstance(obj, LTTextBox):
            x = obj.bbox[0] / page_width
            y = obj.bbox[3] / page_height
            # get text box
            text = obj.get_text()
            numbers_of_signatures = check_signatures(text, x, y, drawn_x, drawn_y,
                                                     numbers_of_signatures)
    return numbers_of_signatures


# A function to create all the info of the path (using PDF scanning and openCV)
# Creates the following:
#   Kornit P/N
#   Drawing number
#   Revision
#   Signatures per page
#   Signature date
#   Missing balloons in assemblies
#   A check if there are too many / too few files for the part number
def get_info(path, pop_path, log, balloons_check, drawing_num, issues):
    # Get the number of balloons in assembly
    max_balloons = number_of_balloons(path)
    # Create a list of balloon numbers from 1-max_balloons
    missing_balloons = list(range(max_balloons + 1))
    missing_balloons.pop(0)
    # Get scanned page layout
    layout = get_pages_obj(path)

    # Initial variables value assignment
    rev_y = 0
    rev = ''
    part_number = None
    drawing_number = None
    signature = False
    date = False
    numbers_of_signatures = 0
    signature_per_page = []

    # Detect circles using openCV in first page of drawing
    if balloons_check == 1:
        [center_x, center_y] = openCV_circles.circles_centers(path, 0, pop_path)
    # For first page of drawing
    for obj in layout[0]:
        page_height = layout[0].height
        page_width = layout[0].width
        if isinstance(obj, LTTextBox):
            # Calibrating x and y coordinates so that they are relative coordinates (from 0 to 1)
            # and do not depend on sheet size
            x = obj.bbox[0] / page_width
            y = obj.bbox[3] / page_height
            # get text box
            text = obj.get_text()
            # If got to the lower right part of the page
            if x > 0.6 and y < 0.2:
                part_number = get_pn(text, part_number)
                drawing_number = get_drawing_number(text, drawing_number)
                if x > 0.85:
                    rev, rev_y = get_rev(text, y, rev_y, rev)
            # Check "Drawn by" location for use in checking signatures and date
            if re.findall("Drawn", text) != [] and y < 0.1:
                x_drawn = x
                y_drawn = y
                # Need to start looking for signature
                signature = True
            # Check for signature, increase counter by amount of signatures in text box
            if signature is True and x_drawn is not None:
                numbers_of_signatures = check_signatures(text, x, y, x_drawn, y_drawn,
                                                         numbers_of_signatures)
            # Check if signature date exists (True if there is a signature date, False if not)
            if date is False:
                date = check_date(text, x, y)
            if balloons_check == 1:
                present_balloon = check_missing_circles(x, y, center_x, center_y, text)
                # check if text is a number in a balloon
                if present_balloon is not None:
                    # Get index of balloon in the missing balloons list
                    index = [i for i, x in enumerate(missing_balloons) if x == present_balloon]
                    # Remove balloon from missing balloons list (because it exists in the drawing)
                    len(index) > 0 and missing_balloons.pop(index[0])
    # Create a list of number of signatures for each page (needs 3 each page)
    if numbers_of_signatures < 3:
        numbers_of_signatures = check_again_signatures(layout[0], page_height, page_width, x_drawn, y_drawn)
    signature_per_page.append(numbers_of_signatures)
    # Last candidate to survive is the true revision
    true_rev = ''.join(rev)

    # For page 2 and onwards, do everything again
    for page_num in range(len(layout[1:])):
        numbers_of_signatures = 0
        page_height = layout[page_num + 1].height
        page_width = layout[page_num + 1].width
        if balloons_check == 1:
            [center_x, center_y] = openCV_circles.circles_centers(path, page_num + 1, pop_path)
        page_layout = layout[page_num + 1]
        for obj in page_layout:
            if isinstance(obj, LTTextBox):
                x = obj.bbox[0] / page_width
                y = obj.bbox[3] / page_height
                text = obj.get_text()
                if re.findall("Drawn", text) != [] and y < 0.1:
                    x_drawn = x
                    y_drawn = y
                    signature = True
                if signature is True and x_drawn is not None:
                    numbers_of_signatures = check_signatures(text, x, y, x_drawn, y_drawn,
                                                             numbers_of_signatures)
                if date is False:
                    date = check_date(text, x, y)
                if balloons_check == 1:
                    present_balloon = check_missing_circles(x, y, center_x, center_y, text)
                    if present_balloon is not None:
                        # print(present_balloon)
                        index = [i for i, x in enumerate(missing_balloons) if x == present_balloon]
                        len(index) > 0 and missing_balloons.pop(index[0])
        if numbers_of_signatures == 0:
            numbers_of_signatures = check_again_signatures(layout[0], page_height, page_width, x_drawn, y_drawn)
        signature_per_page.append(numbers_of_signatures)

    # Print numbers of missing balloons, otherwise print that all balloons exists
    if balloons_check == 1:
        if missing_balloons != []:
            log.insert(tk.END, f"{drawing_num}: Missing balloons: {missing_balloons}\n")
            log.yview(tk.END)
            issues.append(f"Missing balloons: {missing_balloons}")
        # else:
            # log.insert(tk.END, "No missing balloons! Hurray!\n")
    # Return all info of PDF
    return part_number, true_rev, drawing_number, signature_per_page, date
